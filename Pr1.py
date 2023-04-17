import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook , Workbook

ws=load_workbook("Files/x^2-50(sinx)^2.xlsx")["Sheet1"]
max_row=ws.max_row
# print(max_row)

data=[]
for i in range(1,max_row+1):
    data.append([ws[f"A{i}"].value ,ws[f"B{i}"].value ])

data=np.array(data)
# print(data)
x=data[:,0]
y=data[:,1]

dx=x[1:] - x[:-1]
dy=y[1:] - y[:-1]
## print( dx.shape , dy.shape)

###################
##### option1 #####
###################

ydot=dy/dx
## ydot=y.copy()
## ydot[1:]=dy/dx
## ydot[0]=ydot[1]

dydot=ydot[1:] - ydot[:-1]
ydoubledot=dydot/dx[1:]

###################
##### option2 #####
###################

delta=y[2:]+y[:-2] - 2*y[1:-1]
ydoubledot2=delta/dx[1:]**2

plt.plot(x , y ,label="y=x^2-50(sinx)^2.xlsx")
plt.plot(x[1:] , ydot ,label="y dot =2x-50(sin2x).xlsx")
plt.plot(x[2:] , ydoubledot ,label="y double dot =2-100(cos2x).xlsx")
plt.plot(x[2:] , ydoubledot2 ,label="(y double dot)2 =2-100(cos2x).xlsx")
plt.xlabel("X")
plt.ylabel("Y")
plt.legend()
plt.show()

wb_save=Workbook()
ws_save=wb_save.active
i=1
for x_i in x:
    ws_save[f"A{i}"]=x_i
    if i<max_row:
        ws_save[f"B{i}"]=ydot[i-1]
    if i<max_row-1:
        ws_save[f"C{i}"]=ydoubledot[i-1]
        ws_save[f"D{i}"]=ydoubledot2[i-1]
    i+=1
wb_save.save("Files/Derivative.xlsx")